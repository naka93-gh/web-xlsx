"""実ファイル fixture を openpyxl で生成する。

これらの .xlsx は web-xlsx 自身の writer を通さない「独立した生成元」による本物の
OOXML ファイル。Excel と同様に [Content_Types].xml・docProps・完全な名前空間・
numFmts・count 付き sharedStrings を含み、自前 buildXlsx には無い実ファイル構造で
parser を検証する。

再生成: python3 -m venv .venv && .venv/bin/pip install openpyxl && \
        .venv/bin/python tests/fixtures/generate.py
出力は決定的（タイムスタンプ固定）なので、再生成しても差分は出ない。
"""

import re
import zipfile
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

OUT = Path(__file__).parent
# 決定的な出力にするため作成/更新日時を固定する
FIXED = datetime(2024, 1, 1, 0, 0, 0)
FIXED_ISO = FIXED.strftime("%Y-%m-%dT%H:%M:%SZ")


def _freeze_modified(path: Path) -> None:
    """openpyxl は保存時に modified を実時刻で上書きする(writer/excel.py)。

    生成時刻が漏れず再生成で差分も出ないよう、固定値へ書き戻す。
    """
    repl = f'<dcterms:modified xsi:type="dcterms:W3CDTF">{FIXED_ISO}</dcterms:modified>'
    with zipfile.ZipFile(path) as z:
        items = [(i, z.read(i.filename)) for i in z.infolist()]
    with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
        for info, data in items:
            if info.filename == "docProps/core.xml":
                data = re.sub(
                    rb"<dcterms:modified[^>]*>[^<]*</dcterms:modified>",
                    repl.encode(),
                    data,
                )
            # zip エントリ日時も実生成時刻が残るため、固定値(zip 最小の 1980-01-01)にする
            info.date_time = (1980, 1, 1, 0, 0, 0)
            z.writestr(info, data)


def finalize(wb: Workbook, name: str) -> None:
    # creator は openpyxl 既定の "openpyxl"（個人名は入らない）のまま残す
    wb.properties.created = FIXED
    wb.properties.modified = FIXED
    path = OUT / name
    wb.save(path)
    _freeze_modified(path)


def employees() -> None:
    """主軸: 型付き取込。文字列(日本語含む)・整数・日付・真偽・大きい数値。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "社員"
    ws.append(["名前", "年齢", "入社日", "在籍", "給与"])
    rows = [
        ("田中太郎", 30, datetime(2020, 4, 1), True, 4500000),
        ("佐藤花子", 25, datetime(2021, 10, 15), False, 3800000),
        ("John Smith", 41, datetime(2019, 1, 20), True, 5200000),
    ]
    for r in rows:
        ws.append(r)
    # 入社日列を日付書式に（numFmt 由来の日付判定を検証させる）
    for cell in ws["C"][1:]:
        cell.number_format = "yyyy-mm-dd"
    finalize(wb, "employees.xlsx")


def multisheet() -> None:
    """複数シート: 名前/index でのシート選択を検証。"""
    wb = Workbook()
    sales = wb.active
    sales.title = "売上"
    sales.append(["月", "金額"])
    sales.append(["1月", 100])
    sales.append(["2月", 200])

    cost = wb.create_sheet("費用")
    cost.append(["月", "金額"])
    cost.append(["1月", 50])

    summary = wb.create_sheet("Summary")
    summary.append(["項目", "値"])
    summary.append(["利益", 250])
    finalize(wb, "multisheet.xlsx")


def edge() -> None:
    """エッジ: テキスト格納ID(大整数・先頭ゼロ)・前後空白・空セル・浮動小数・負数。

    19桁IDや先頭ゼロは Excel でも数値だと桁落ち/欠落する。実運用ではテキスト列に
    入れて保持するのが定石で、string 指定で raw 文字列を読めるかを検証する。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Edge"
    ws.append(["ID", "名前", "メモ", "数値"])
    # ID はテキスト格納（数値だと桁落ち/先頭ゼロ欠落するため）
    ws.append(["1234567890123456789", "  前後  ", None, 3.14])
    ws.append(["0042", "通常", "テキスト", -7])
    finalize(wb, "edge.xlsx")


if __name__ == "__main__":
    employees()
    multisheet()
    edge()
    print("generated:", *(p.name for p in sorted(OUT.glob("*.xlsx"))))
